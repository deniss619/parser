import datetime
import time

from selenium import webdriver
import openpyxl
from openpyxl.utils import get_column_letter
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from platform import python_version


class Excel:
    def __init__(self, data):
        self.data = data

    def make_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.append(("Дата", "Курс", "Изменение", "Дата", "Курс", "Изменение", "Euro/Dollar"))
        for i in self.data:
            ws.append(i)
        for i in range(2, len(self.data) + 2):
            for j in [2, 3, 5, 6]:
                ws.cell(i, j).number_format = u'_-* # ##0.00 ₽_-;-* # ##0.00 ₽_-;'
            ws.cell(i, 7).number_format = u'_-* # ##0.00_-;-* # ##0.00_-;'
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        wb.save("ex1.xlsx")


class Email:
    def __init__(self, server, user, password, recipients, subject, text):
        self.server = server
        self.user = user
        self.password = password
        self.recipients = recipients
        self.subject = subject
        self.text = text
        self.msg = MIMEMultipart('alternative')

    def generate_mail(self):
        self.msg['Subject'] = self.subject
        self.msg['From'] = 'Python script <' + self.user + '>'
        self.msg['To'] = ', '.join(self.recipients)
        self.msg['Reply-To'] = self.user
        self.msg['Return-Path'] = self.user
        self.msg['X-Mailer'] = 'Python/' + (python_version())
        part_text = MIMEText(self.text, 'plain')
        self.msg.attach(part_text)

    def add_file(self, filepath):
        basename = os.path.basename(filepath)
        filesize = os.path.getsize(filepath)
        part_file = MIMEBase('application', 'octet-stream; name="{}"'.format(basename))
        part_file.set_payload(open(filepath, "rb").read())
        part_file.add_header('Content-Description', basename)
        part_file.add_header('Content-Disposition', 'attachment; filename="{}"; size={}'.format(basename, filesize))
        encoders.encode_base64(part_file)
        self.msg.attach(part_file)

    def add_html(self):
        html = '<html><head></head><body><p>' + self.text + '</p></body></html>'
        part_html = MIMEText(html, 'html')
        self.msg.attach(part_html)

    def send_mail(self):
        mail = smtplib.SMTP_SSL(self.server)
        mail.login(self.user, self.password)
        mail.sendmail(self.user, self.recipients, self.msg.as_string())
        mail.quit()


def get_end(l):
    l = l % 100
    if (l > 20):
        return get_end(l % 10)

    if (l == 1):
        return 'а'
    elif (2 <= l <= 4):
        return 'и'
    else:
        return ''


def create_driver():
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--start-maximized")
    return webdriver.Chrome(options=chrome_options)


class Parser:
    def __init__(self):
        self.driver = create_driver()

    def go_to_necessary_page(self):
        self.driver.get("https://www.moex.com/")
        self.driver.find_element_by_class_name("js-menu-dropdown-button").click()
        self.driver.find_element_by_xpath(
            "//div[@class='header-menu-dropdown__inner']/div[@class='menu-block'][1]/div[@class='item'][2]/a").click()
        time.sleep(1)
        self.driver.find_element_by_xpath("//div[@class='disclaimer__buttons']/a").click()
        self.driver.find_element_by_xpath("//a[@class='sidebar-list__item childs-closed'][6]").click()
        return self.driver.current_url

    def parse(self, url):
        self.driver.get(url)
        time.sleep(1)
        return self.get_data()

    def get_data(self):
        table = self.driver.find_element_by_class_name("tablels")
        data = []
        breaker = False
        for i in range(1, table.text.count('\n') // 2):
            for j in [1, 0]:
                str1 = table.find_element_by_xpath("//tbody/tr[@class='tr%s'][%s]/td[1]" % (j, i)).text
                if int(str1[3:5]) == datetime.datetime.now().month:
                    exchangeRate1 = float(
                        table.find_element_by_xpath("//tbody/tr[@class='tr%s'][%s]/td[2]" % (j, i)).text.replace(',',
                                                                                                                 '.'))
                    exchangeRate2 = float(table.find_element_by_xpath(
                        "//tbody/tr[@class='tr%s'][%s]/td[2]" % (int(not bool(j)), i + 1)).text.replace(',', '.'))
                    diff = round(exchangeRate1 - exchangeRate2, 4)
                    data.append([str1, exchangeRate1, diff])
                else:
                    breaker = True
            if j == 0 and breaker:
                break
        return data

    def main(self):
        url = self.go_to_necessary_page()
        dollar = self.parse(url[:-7] + 'USD_RUB')
        euro = self.parse(url[:-7] + 'EUR_RUB')
        self.driver.close()
        return dollar, euro


if __name__ == '__main__':

    dollar, euro = Parser().main()
    division = [[round(float(euro[i][1]) / float(dollar[i][1]), 4)] for i in range(len(euro))]

    data = [dollar[i] + euro[i] + division[i] for i in range(len(dollar))]
    print(data)

    excel = Excel(data)
    excel.make_excel()
    email = Email('smtp.gmail.com', 'email@gmail.com', "password", ['email2@gmail.com'], 'Отчет',
                  f"В файле {len(data)} строк" + get_end(len(data))) # заменить email@gmail.com на почту отправителя,
    # заменить password на пароль от почты отправитлея, заменить ['email2@gmail.com'] на список почт получателей
    email.generate_mail()
    email.add_file("ex1.xlsx")
    email.send_mail()
